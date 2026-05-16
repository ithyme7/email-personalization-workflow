from __future__ import annotations

from collections import Counter
from pathlib import Path
import shutil
from typing import Any
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import OUTPUT_COLUMNS
from sendability import evaluate_sendability


CLIENT_REVIEW_COLUMNS = [
    "status",
    "sendability_decision",
    "sendability_score",
    "sendability_reasons",
    "human_decision",
    "edited_line",
    "edit_reason_category",
    "edit_notes",
    "company",
    "person",
    "personalized_line",
    "template_preview",
    "friction_type",
    "surface_checked",
    "conversion_outcome",
    "product_surface_type",
    "research_priority",
    "angle_gate_decision",
    "app_check_status",
    "recommended_manual_check",
    "visual_flags",
    "visual_confidence",
    "visual_confidence_score",
    "visual_confidence_reasons",
    "evidence_found",
    "screenshots",
    "shareable_screenshots",
    "trace_files",
    "ux_validator_findings",
    "advanced_detector_flags",
    "dead_link_checks",
    "needs_manual_review",
    "quality_flags",
    "reviewer_notes",
    "role",
    "website",
    "source_urls",
    "llm_calls",
    "estimated_input_tokens",
    "estimated_output_tokens",
]

CLIENT_RESEARCH_COLUMNS = [
    "company",
    "person",
    "role",
    "website",
    "tone profile",
    "model provider",
    "model name",
    "llm calls",
    "estimated input tokens",
    "estimated output tokens",
    "linkedin observation",
    "app flow observation",
    "product surface type",
    "research priority",
    "app review complaints",
    "app check status",
    "recommended manual check",
    "template preview",
    "app store summary",
    "visual observations",
    "visual flags",
    "visual confidence",
    "visual confidence reasons",
    "screenshot paths",
    "shareable screenshots",
    "trace files",
    "ux validator findings",
    "advanced detector flags",
    "dead link checks",
    "friction type",
    "surface checked",
    "conversion outcome",
    "angle gate decision",
    "angle gate notes",
    "blocked angles",
    "angle priority",
    "blog used",
    "why this angle",
    "evidence found",
    "evidence strength score",
    "personalization quality score",
    "source urls",
    "raw reviewer notes",
]

SUMMARY_COLUMNS = ["metric", "value"]

DARK_BG = "0F172A"
DARK_PANEL = "111827"
DARK_PANEL_2 = "1E293B"
DARK_BORDER = "334155"
TEXT_MAIN = "F8FAFC"
TEXT_MUTED = "CBD5E1"
ACCENT_BLUE = "38BDF8"
ACCENT_GREEN = "22C55E"
ACCENT_ORANGE = "FB923C"
ACCENT_RED = "F87171"
ACCENT_PURPLE = "A78BFA"
ACCENT_YELLOW = "FACC15"


def export_rows(rows: list[dict[str, Any]], output_path: str | Path) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    normalized_rows = []
    for row in rows:
        normalized_rows.append({column: row.get(column, "") for column in OUTPUT_COLUMNS})

    df = pd.DataFrame(normalized_rows, columns=OUTPUT_COLUMNS)
    if path.suffix.lower() == ".xlsx":
        df.to_excel(path, index=False)
    else:
        df.to_csv(path, index=False, encoding="utf-8")


def _shorten(value: Any, limit: int = 420) -> str:
    text = str(value or "").replace("\r", " ").replace("\n", " ")
    replacements = {
        "â\x80\x8e": "",
        "â\x80\x91": "-",
        "â\x80\x93": "-",
        "â\x80\x94": "-",
        "â\x80\x99": "'",
        "â\x80\x9c": '"',
        "â\x80\x9d": '"',
        "â\x80\xa6": "...",
        "Â·": "-",
        "Â": "",
        "—": "-",
        "–": "-",
        "’": "'",
        "“": '"',
        "”": '"',
        "·": "-",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    text = " ".join(text.split())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def _yes_no(value: Any) -> str:
    if isinstance(value, bool):
        return "yes" if value else "no"
    text = str(value or "").strip().lower()
    if text in {"true", "waar", "yes", "1"}:
        return "yes"
    if text in {"false", "onwaar", "no", "0"}:
        return "no"
    return str(value or "")


def _clean_notes(value: Any) -> str:
    notes = str(value or "")
    noisy = [
        "LinkedIn URL supplied, but no manual LinkedIn observation provided. LinkedIn was not scraped.",
        "No manual app/onboarding walkthrough observation provided.",
    ]
    for item in noisy:
        notes = notes.replace(item, "")
    notes = notes.replace(" |  | ", " | ").strip(" |")
    return _shorten(notes, 520)


def _cell_lines(value: Any, limit: int = 900) -> str:
    text = _shorten(value, limit)
    return text.replace(" | ", "\n").replace("; ", "\n")


def _is_manual_review(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"true", "waar", "yes", "1", "review"}


def _status_for(row: dict[str, Any], personalized_line: str) -> str:
    flags = str(row.get("quality_flags", "")).lower()
    if not personalized_line.strip() or personalized_line.strip().startswith("["):
        return "Research only"
    if "ai_generation_unavailable" in flags:
        return "Research only"
    if _is_manual_review(row.get("needs_manual_review")):
        return "Review"
    return "Ready"


def _client_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    review_rows: list[dict[str, Any]] = []
    research_rows: list[dict[str, Any]] = []
    for row in rows:
        evidence = row.get("evidence_used_for_copy") or row.get("evidence_points", "")
        personalized_line = row.get("opening_line", "")
        notes = _clean_notes(row.get("reviewer_notes", ""))
        if not personalized_line and "ai_generation_unavailable" in str(row.get("quality_flags", "")):
            personalized_line = "[research only, AI writing not available]"
        status = _status_for(row, personalized_line)
        review_row = {
            "status": status,
            "company": row.get("company_name", ""),
            "person": row.get("recipient_name", ""),
            "personalized_line": personalized_line,
            "template_preview": _cell_lines(row.get("template_preview", ""), 520),
            "friction_type": row.get("friction_type", ""),
            "surface_checked": row.get("surface_checked", ""),
            "conversion_outcome": row.get("conversion_outcome", ""),
            "product_surface_type": row.get("product_surface_type", ""),
            "research_priority": row.get("research_priority", ""),
            "angle_gate_decision": row.get("angle_gate_decision", ""),
            "app_check_status": row.get("app_check_status", ""),
            "recommended_manual_check": _cell_lines(row.get("recommended_manual_check", ""), 520),
            "visual_flags": _cell_lines(row.get("visual_quality_flags", ""), 240),
            "visual_confidence": row.get("visual_confidence", ""),
            "visual_confidence_score": row.get("visual_confidence_score", ""),
            "visual_confidence_reasons": _cell_lines(row.get("visual_confidence_reasons", ""), 360),
            "evidence_found": _cell_lines(evidence, 760),
            "screenshots": _cell_lines(row.get("screenshot_paths", ""), 520),
            "shareable_screenshots": _cell_lines(row.get("shareable_screenshot_files", ""), 520),
            "trace_files": _cell_lines(row.get("trace_files", ""), 520),
            "ux_validator_findings": _cell_lines(row.get("ux_validator_findings", ""), 520),
            "advanced_detector_flags": _cell_lines(row.get("advanced_detector_flags", ""), 240),
            "dead_link_checks": _cell_lines(row.get("dead_link_checks", ""), 420),
            "needs_manual_review": _yes_no(row.get("needs_manual_review", "")),
            "quality_flags": _cell_lines(row.get("quality_flags", ""), 240),
            "reviewer_notes": notes,
            "role": row.get("recipient_role", ""),
            "website": row.get("website_url", ""),
            "source_urls": _cell_lines(row.get("source_urls", ""), 520),
            "llm_calls": row.get("llm_calls", ""),
            "estimated_input_tokens": row.get("estimated_input_tokens", ""),
            "estimated_output_tokens": row.get("estimated_output_tokens", ""),
        }
        review_row.update(evaluate_sendability(review_row))
        review_row.setdefault("human_decision", "unreviewed")
        review_row.setdefault("edited_line", "")
        review_row.setdefault("edit_reason_category", "not_reviewed")
        review_row.setdefault("edit_notes", "")
        review_rows.append(review_row)
        research_rows.append(
            {
                "company": row.get("company_name", ""),
                "person": row.get("recipient_name", ""),
                "role": row.get("recipient_role", ""),
                "website": row.get("website_url", ""),
                "tone profile": row.get("tone_profile", ""),
                "model provider": row.get("model_provider", ""),
                "model name": row.get("model_name", ""),
                "llm calls": row.get("llm_calls", ""),
                "estimated input tokens": row.get("estimated_input_tokens", ""),
                "estimated output tokens": row.get("estimated_output_tokens", ""),
                "linkedin observation": row.get("linkedin_observation", ""),
                "app flow observation": row.get("app_flow_observation", ""),
                "product surface type": row.get("product_surface_type", ""),
                "research priority": row.get("research_priority", ""),
                "app review complaints": _shorten(row.get("app_review_complaints", ""), 900),
                "app check status": row.get("app_check_status", ""),
                "recommended manual check": _shorten(row.get("recommended_manual_check", ""), 900),
                "template preview": _shorten(row.get("template_preview", ""), 900),
                "app store summary": _shorten(row.get("app_store_summary", ""), 900),
                "visual observations": _shorten(row.get("visual_observations", ""), 900),
                "visual flags": _shorten(row.get("visual_quality_flags", ""), 500),
                "visual confidence": _shorten(
                    f"{row.get('visual_confidence', '')} ({row.get('visual_confidence_score', '')}/100)", 80
                ),
                "visual confidence reasons": _shorten(row.get("visual_confidence_reasons", ""), 900),
                "screenshot paths": _shorten(row.get("screenshot_paths", ""), 900),
                "shareable screenshots": _shorten(row.get("shareable_screenshot_files", ""), 900),
                "trace files": _shorten(row.get("trace_files", ""), 900),
                "ux validator findings": _shorten(row.get("ux_validator_findings", ""), 900),
                "advanced detector flags": _shorten(row.get("advanced_detector_flags", ""), 500),
                "dead link checks": _shorten(row.get("dead_link_checks", ""), 900),
                "friction type": row.get("friction_type", ""),
                "surface checked": row.get("surface_checked", ""),
                "conversion outcome": row.get("conversion_outcome", ""),
                "angle gate decision": row.get("angle_gate_decision", ""),
                "angle gate notes": _shorten(row.get("angle_gate_notes", ""), 500),
                "blocked angles": _shorten(row.get("blocked_angles", ""), 900),
                "angle priority": row.get("angle_priority", ""),
                "blog used": row.get("blog_used", ""),
                "why this angle": _shorten(row.get("why_this_angle", ""), 500),
                "evidence found": _shorten(evidence, 900),
                "evidence strength score": row.get("evidence_strength_score", ""),
                "personalization quality score": row.get("personalization_quality_score", ""),
                "source urls": _shorten(row.get("source_urls", ""), 500),
                "raw reviewer notes": notes,
            }
        )
    return review_rows, research_rows


def _style_sheet(ws, widths: list[int], row_height: int = 54, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = row_height
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = freeze


def _style_review_status(ws) -> None:
    status_fills = {
        "Ready": PatternFill("solid", fgColor="D9EAD3"),
        "Review": PatternFill("solid", fgColor="FCE4D6"),
        "Research only": PatternFill("solid", fgColor="E7E6E6"),
    }
    sendability_fills = {
        "Send": PatternFill("solid", fgColor="D9EAD3"),
        "Edit": PatternFill("solid", fgColor="FCE4D6"),
        "Reject": PatternFill("solid", fgColor="F4CCCC"),
    }
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        fill = status_fills.get(str(cell.value or ""))
        if fill:
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        fill = sendability_fills.get(str(cell.value or ""))
        if fill:
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _summary_rows(review_rows: list[dict[str, Any]], research_rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    unique_companies = len({str(row.get("company", "")).strip().lower() for row in review_rows if row.get("company")})
    research_only = sum(1 for row in review_rows if row.get("status") == "Research only")
    review_needed = sum(1 for row in review_rows if _is_manual_review(row.get("needs_manual_review")))
    ready = sum(1 for row in review_rows if row.get("status") == "Ready")
    sendable = sum(1 for row in review_rows if row.get("sendability_decision") == "Send")
    edit_needed = sum(1 for row in review_rows if row.get("sendability_decision") == "Edit")
    reject_needed = sum(1 for row in review_rows if row.get("sendability_decision") == "Reject")
    generated = sum(
        1
        for row in review_rows
        if str(row.get("personalized_line", "")).strip()
        and not str(row.get("personalized_line", "")).strip().startswith("[")
    )
    sources = sum(1 for row in research_rows if str(row.get("source urls", "")).strip())
    high_visual = sum(1 for row in review_rows if str(row.get("visual_confidence", "")).lower() == "high")
    medium_visual = sum(1 for row in review_rows if str(row.get("visual_confidence", "")).lower() == "medium")
    low_visual = sum(1 for row in review_rows if str(row.get("visual_confidence", "")).lower() == "low")
    app_checks = sum(
        1 for row in review_rows if "recommended" in str(row.get("app_check_status", "")).lower()
    )
    tone_profiles = sorted({str(row.get("tone profile", "")).strip() for row in research_rows if row.get("tone profile")})
    model_names = sorted({str(row.get("model name", "")).strip() for row in research_rows if row.get("model name")})
    return [
        {"metric": "Total contact rows", "value": str(len(review_rows))},
        {"metric": "Unique companies", "value": str(unique_companies)},
        {"metric": "Rows with personalized line", "value": str(generated)},
        {"metric": "Rows ready to send", "value": str(ready)},
        {"metric": "Sendability: Send", "value": str(sendable)},
        {"metric": "Sendability: Edit", "value": str(edit_needed)},
        {"metric": "Sendability: Reject", "value": str(reject_needed)},
        {"metric": "Rows needing review", "value": str(review_needed)},
        {"metric": "Research-only rows", "value": str(research_only)},
        {"metric": "Rows with source URLs", "value": str(sources)},
        {"metric": "High visual-confidence rows", "value": str(high_visual)},
        {"metric": "Medium visual-confidence rows", "value": str(medium_visual)},
        {"metric": "Low visual-confidence rows", "value": str(low_visual)},
        {"metric": "Rows where app walkthrough is recommended", "value": str(app_checks)},
        {"metric": "Tone profile", "value": ", ".join(tone_profiles) or "not set"},
        {"metric": "Model", "value": ", ".join(model_names) or "not set"},
        {
            "metric": "How to use",
            "value": "Start in the Review tab. Edit personalized_line if needed, then check Research Details for longer evidence.",
        },
    ]


def _write_sheet(ws, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])


def _split_flags(value: Any) -> list[str]:
    text = str(value or "").replace("\n", "|").replace(";", "|")
    return [item.strip() for item in text.split("|") if item.strip()]


def _top_counts(rows: list[dict[str, Any]], key: str, limit: int = 8, split: bool = False) -> list[tuple[str, int]]:
    counter: Counter[str] = Counter()
    for row in rows:
        value = row.get(key, "")
        if split:
            counter.update(_split_flags(value))
        else:
            text = str(value or "").strip()
            if text:
                counter[text] += 1
    return counter.most_common(limit)


def _numeric_values(rows: list[dict[str, Any]], key: str) -> list[float]:
    values: list[float] = []
    for row in rows:
        try:
            value = float(row.get(key, ""))
        except (TypeError, ValueError):
            continue
        if value > 0:
            values.append(value)
    return values


def _dashboard_fill(ws, start_row: int = 1, end_row: int = 54, start_col: int = 1, end_col: int = 12) -> None:
    fill = PatternFill("solid", fgColor=DARK_BG)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.fill = fill
            cell.font = Font(color=TEXT_MAIN)
            cell.border = Border()


def _merge_panel(ws, cell_range: str, fill: str = DARK_PANEL) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = Border(
                left=Side(style="thin", color=DARK_BORDER),
                right=Side(style="thin", color=DARK_BORDER),
                top=Side(style="thin", color=DARK_BORDER),
                bottom=Side(style="thin", color=DARK_BORDER),
            )


def _kpi_card(ws, cell_range: str, title: str, value: str, subtitle: str, accent: str) -> None:
    _merge_panel(ws, cell_range, DARK_PANEL)
    top_left = cell_range.split(":", 1)[0]
    row = ws[top_left].row
    col = ws[top_left].column
    ws.cell(row, col).value = title.upper()
    ws.cell(row, col).font = Font(color=TEXT_MUTED, bold=True, size=9)
    ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="top")
    ws.cell(row + 1, col).value = value
    ws.cell(row + 1, col).font = Font(color=accent, bold=True, size=22)
    ws.cell(row + 1, col).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row + 2, col).value = subtitle
    ws.cell(row + 2, col).font = Font(color=TEXT_MUTED, size=9)
    ws.cell(row + 2, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _section_title(ws, cell: str, title: str, subtitle: str = "") -> None:
    ws[cell].value = title
    ws[cell].font = Font(color=TEXT_MAIN, bold=True, size=14)
    ws[cell].alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        row = ws[cell].row + 1
        col = ws[cell].column
        ws.cell(row, col).value = subtitle
        ws.cell(row, col).font = Font(color=TEXT_MUTED, size=9)
        ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _write_dashboard_table(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    rows: list[tuple[str, int]],
    label_header: str = "label",
    value_header: str = "count",
) -> tuple[int, int]:
    ws.cell(start_row, start_col).value = title
    ws.cell(start_row, start_col).font = Font(color=TEXT_MAIN, bold=True, size=12)
    ws.cell(start_row + 1, start_col).value = label_header
    ws.cell(start_row + 1, start_col + 1).value = value_header
    for cell in [ws.cell(start_row + 1, start_col), ws.cell(start_row + 1, start_col + 1)]:
        cell.font = Font(color=TEXT_MUTED, bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor=DARK_PANEL_2)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for offset, (label, count) in enumerate(rows, 2):
        row_idx = start_row + offset
        ws.cell(row_idx, start_col).value = label
        ws.cell(row_idx, start_col + 1).value = count
        ws.cell(row_idx, start_col).font = Font(color=TEXT_MAIN, size=9)
        ws.cell(row_idx, start_col + 1).font = Font(color=TEXT_MAIN, bold=True, size=9)
        ws.cell(row_idx, start_col).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row_idx, start_col + 1).alignment = Alignment(horizontal="center", vertical="center")
    return start_row + 1, start_row + max(len(rows), 1) + 1


def _write_metric_bars(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    rows: list[tuple[str, int]],
    label_header: str,
    accent: str,
    max_rows: int = 7,
) -> None:
    rows = rows[:max_rows] or [("No data", 0)]
    max_value = max([count for _, count in rows] + [1])
    ws.cell(start_row, start_col).value = title
    ws.cell(start_row, start_col).font = Font(color=TEXT_MAIN, bold=True, size=12)
    ws.cell(start_row, start_col).alignment = Alignment(horizontal="left", vertical="center")
    headers = [label_header, "rows", "share"]
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + offset)
        cell.value = header
        cell.font = Font(color=TEXT_MUTED, bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor=DARK_PANEL_2)
        cell.alignment = Alignment(horizontal="left" if offset == 0 else "center", vertical="center")
    for idx, (label, count) in enumerate(rows, 2):
        row_idx = start_row + idx
        width = max(1, round((count / max_value) * 18))
        bar = "█" * width
        ws.cell(row_idx, start_col).value = label
        ws.cell(row_idx, start_col + 1).value = count
        ws.cell(row_idx, start_col + 2).value = bar
        ws.cell(row_idx, start_col).font = Font(color=TEXT_MAIN, size=9)
        ws.cell(row_idx, start_col + 1).font = Font(color=TEXT_MAIN, bold=True, size=9)
        ws.cell(row_idx, start_col + 2).font = Font(color=accent, bold=True, size=9)
        ws.cell(row_idx, start_col).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row_idx, start_col + 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_idx, start_col + 2).alignment = Alignment(horizontal="left", vertical="center")


def _style_chart(chart, title: str) -> None:
    chart.title = title
    chart.style = 13
    chart.legend.position = "r"
    try:
        chart.graphical_properties.solidFill = DARK_PANEL
        chart.graphical_properties.line.solidFill = DARK_BORDER
        chart.plot_area.graphicalProperties.solidFill = DARK_PANEL
        chart.plot_area.graphicalProperties.line.solidFill = DARK_BORDER
    except AttributeError:
        pass


def _add_bar_chart(ws, data_min_row: int, data_max_row: int, label_col: int, value_col: int, anchor: str, title: str) -> None:
    if data_max_row <= data_min_row:
        return
    chart = BarChart()
    values = Reference(ws, min_col=value_col, min_row=data_min_row, max_row=data_max_row)
    labels = Reference(ws, min_col=label_col, min_row=data_min_row + 1, max_row=data_max_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(labels)
    chart.type = "bar"
    chart.height = 7
    chart.width = 12
    chart.y_axis.majorGridlines = None
    _style_chart(chart, title)
    ws.add_chart(chart, anchor)


def _add_doughnut_chart(ws, data_min_row: int, data_max_row: int, label_col: int, value_col: int, anchor: str, title: str) -> None:
    if data_max_row <= data_min_row:
        return
    chart = DoughnutChart()
    values = Reference(ws, min_col=value_col, min_row=data_min_row, max_row=data_max_row)
    labels = Reference(ws, min_col=label_col, min_row=data_min_row + 1, max_row=data_max_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(labels)
    chart.holeSize = 58
    chart.height = 7
    chart.width = 9
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    _style_chart(chart, title)
    ws.add_chart(chart, anchor)


def _write_dashboard(ws, review_rows: list[dict[str, Any]], research_rows: list[dict[str, Any]]) -> None:
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"
    for idx, width in enumerate([3, 17, 17, 3, 17, 17, 3, 17, 17, 3, 17, 17], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(1, 55):
        ws.row_dimensions[row].height = 22
    _dashboard_fill(ws)

    total_rows = len(review_rows)
    unique_companies = len({str(row.get("company", "")).strip().lower() for row in review_rows if row.get("company")})
    ready = sum(1 for row in review_rows if row.get("status") == "Ready")
    review = sum(1 for row in review_rows if row.get("status") == "Review")
    research_only = sum(1 for row in review_rows if row.get("status") == "Research only")
    sendable = sum(1 for row in review_rows if row.get("sendability_decision") == "Send")
    edit_needed = sum(1 for row in review_rows if row.get("sendability_decision") == "Edit")
    reject_needed = sum(1 for row in review_rows if row.get("sendability_decision") == "Reject")
    generated = sum(
        1
        for row in review_rows
        if str(row.get("personalized_line", "")).strip()
        and not str(row.get("personalized_line", "")).strip().startswith("[")
    )
    sources = sum(1 for row in research_rows if str(row.get("source urls", "")).strip())
    quality_scores = _numeric_values(research_rows, "personalization quality score")
    avg_quality = round(sum(quality_scores) / len(quality_scores), 1) if quality_scores else 0
    ready_rate = round((ready / total_rows) * 100) if total_rows else 0

    _merge_panel(ws, "B2:L4", DARK_PANEL)
    ws["B2"] = "Email Personalization Batch Dashboard"
    ws["B2"].font = Font(color=TEXT_MAIN, bold=True, size=22)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B3"] = "Quick read on output quality, review workload, evidence strength and visual research confidence."
    ws["B3"].font = Font(color=TEXT_MUTED, size=10)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="top")

    _kpi_card(ws, "B6:C8", "Rows", str(total_rows), f"{unique_companies} unique companies", ACCENT_BLUE)
    _kpi_card(ws, "E6:F8", "Send", str(sendable), "Passes sendability gate", ACCENT_GREEN)
    _kpi_card(ws, "H6:I8", "Edit", str(edit_needed), "Needs light human edit", ACCENT_ORANGE)
    _kpi_card(ws, "K6:L8", "Reject", str(reject_needed), "Do not send yet", ACCENT_RED)
    _kpi_card(ws, "B10:C12", "Generated Lines", str(generated), "Rows with personalization", ACCENT_PURPLE)
    _kpi_card(ws, "E10:F12", "Source Coverage", str(sources), "Rows with source URLs", ACCENT_BLUE)
    _kpi_card(ws, "H10:I12", "Avg QC", str(avg_quality or "-"), "Personalization quality score", ACCENT_YELLOW)
    visual_conf_counts = Counter(str(row.get("visual_confidence", "") or "none").lower() for row in review_rows)
    strongest_visual = max(["high", "medium", "low", "none"], key=lambda key: visual_conf_counts.get(key, 0))
    _kpi_card(ws, "K10:L12", "Visual Signal", strongest_visual.title(), "Most common confidence", ACCENT_GREEN if strongest_visual == "high" else ACCENT_ORANGE)

    status_rows = [(label, count) for label, count in Counter(row.get("status", "") for row in review_rows).items() if label]
    status_order = {"Ready": 0, "Review": 1, "Research only": 2}
    status_rows = sorted(status_rows, key=lambda item: status_order.get(item[0], 9))
    sendability_rows = [
        ("Send", sendable),
        ("Edit", edit_needed),
        ("Reject", reject_needed),
    ]
    sendability_rows = [row for row in sendability_rows if row[1] > 0]
    confidence_rows = [
        ("High", visual_conf_counts.get("high", 0)),
        ("Medium", visual_conf_counts.get("medium", 0)),
        ("Low", visual_conf_counts.get("low", 0)),
        ("None", visual_conf_counts.get("none", 0) + visual_conf_counts.get("", 0)),
    ]
    confidence_rows = [row for row in confidence_rows if row[1] > 0]
    friction_rows = _top_counts(review_rows, "friction_type", limit=8)
    visual_flag_rows = _top_counts(review_rows, "visual_flags", limit=8, split=True)
    outcome_rows = _top_counts(review_rows, "conversion_outcome", limit=8)
    review_reason_rows = _top_counts(review_rows, "quality_flags", limit=8, split=True)
    sendability_reason_rows = _top_counts(review_rows, "sendability_reasons", limit=8, split=True)

    _section_title(ws, "B15", "Sendability Gate", "Rows are separated into send, edit, and reject before delivery.")
    _write_metric_bars(ws, 17, 2, "Sendability distribution", sendability_rows or [("No rows", 0)], "decision", ACCENT_GREEN)

    _section_title(ws, "H15", "Visual Confidence", "Reliability of automated visual findings.")
    _write_metric_bars(ws, 17, 8, "Visual confidence", confidence_rows or [("None", 0)], "confidence", ACCENT_BLUE)

    _section_title(ws, "B28", "Top Friction Types", "The angle gate should prefer current conversion or UX friction.")
    _write_metric_bars(ws, 30, 2, "Friction type ranking", friction_rows or [("No friction selected", 0)], "friction type", ACCENT_ORANGE)

    _section_title(ws, "H28", "Outcome Coverage", "What the lines connect to: activation, conversion, drop-off, bookings, trust.")
    _write_metric_bars(ws, 30, 8, "Conversion outcomes", outcome_rows or [("No outcome", 0)], "outcome", ACCENT_PURPLE)

    _section_title(ws, "B43", "Visual Flags", "Automated screen checks. High/medium confidence findings are more useful.")
    _write_metric_bars(ws, 45, 2, "Top visual flags", visual_flag_rows or [("No visual flags", 0)], "flag", ACCENT_YELLOW)

    _section_title(ws, "H43", "Review Queue", "Main reasons rows still need human attention.")
    _write_metric_bars(ws, 45, 8, "Sendability reasons", sendability_reason_rows or review_reason_rows or [("No quality flags", 0)], "reason", ACCENT_RED)

    for row in ws.iter_rows(min_row=15, max_row=54, min_col=2, max_col=12):
        for cell in row:
            if cell.value is not None and cell.fill.fgColor.rgb in {None, "00000000"}:
                cell.fill = PatternFill("solid", fgColor=DARK_BG)
            cell.border = Border(bottom=Side(style="thin", color="1E293B"))
            if cell.column in {3, 9} and isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_readable_xlsx(rows: list[dict[str, Any]], output_path: Path) -> None:
    review_rows, research_rows = _client_rows(rows)
    wb = Workbook()
    dashboard = wb.active
    _write_dashboard(dashboard, review_rows, research_rows)

    ws = wb.create_sheet("Review")
    _write_sheet(ws, CLIENT_REVIEW_COLUMNS, review_rows)
    _style_sheet(
        ws,
        [14, 14, 12, 48, 16, 70, 22, 42, 22, 20, 72, 78, 24, 24, 24, 28, 26, 64, 28, 18, 72, 52, 16, 26, 56, 24, 30, 48],
        row_height=88,
        freeze="K2",
    )
    _style_review_status(ws)

    details = wb.create_sheet("Research Details")
    _write_sheet(details, CLIENT_RESEARCH_COLUMNS, research_rows)
    _style_sheet(
        details,
        [22, 20, 24, 30, 20, 20, 28, 42, 48, 26, 70, 78, 78, 78, 36, 20, 78, 78, 24, 24, 24, 28, 52, 78, 16, 14, 52, 78, 18, 22, 52, 60],
        row_height=92,
    )

    summary = wb.create_sheet("Summary")
    _write_sheet(summary, SUMMARY_COLUMNS, _summary_rows(review_rows, research_rows))
    _style_sheet(summary, [28, 90], row_height=38)
    summary.sheet_properties.tabColor = "70AD47"
    dashboard.sheet_properties.tabColor = "111827"
    ws.sheet_properties.tabColor = "5B9BD5"
    details.sheet_properties.tabColor = "A5A5A5"
    wb.active = 0

    wb.save(output_path)


def _split_file_list(value: Any) -> list[Path]:
    paths: list[Path] = []
    for item in str(value or "").replace("\n", "|").split("|"):
        item = item.strip()
        if not item:
            continue
        path = Path(item)
        if path.exists() and path.is_file():
            paths.append(path)
    return paths


def _copy_assets_for_delivery(rows: list[dict[str, Any]], output_path: Path) -> Path | None:
    assets_root = output_path.with_name(f"{output_path.stem}_assets")
    screenshots_dir = assets_root / "screenshots"
    traces_dir = assets_root / "traces"
    copied_files: list[Path] = []

    for row in rows:
        shareable_screenshots: list[str] = []
        shareable_traces: list[str] = []
        for source in _split_file_list(row.get("screenshot_paths", "")):
            screenshots_dir.mkdir(parents=True, exist_ok=True)
            target = screenshots_dir / source.name
            if not target.exists():
                shutil.copy2(source, target)
            copied_files.append(target)
            shareable_screenshots.append(str(target.relative_to(output_path.parent)))
        for source in _split_file_list(row.get("trace_files", "")):
            traces_dir.mkdir(parents=True, exist_ok=True)
            target = traces_dir / source.name
            if not target.exists():
                shutil.copy2(source, target)
            copied_files.append(target)
            shareable_traces.append(str(target.relative_to(output_path.parent)))
        if shareable_screenshots:
            row["shareable_screenshot_files"] = " | ".join(dict.fromkeys(shareable_screenshots))
        if shareable_traces:
            row["trace_files"] = " | ".join(dict.fromkeys(shareable_traces))

    if not copied_files:
        return None

    package_path = output_path.with_name(f"{output_path.stem}_delivery_package.zip")
    with zipfile.ZipFile(package_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        if output_path.exists():
            archive.write(output_path, output_path.name)
        for file_path in copied_files:
            archive.write(file_path, file_path.relative_to(output_path.parent))
    return package_path


def export_client_batch_rows(rows: list[dict[str, Any]], output_path: str | Path) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    _copy_assets_for_delivery(rows, path)
    if path.suffix.lower() == ".xlsx":
        _write_readable_xlsx(rows, path)
        _copy_assets_for_delivery(rows, path)
    else:
        compact_rows, _ = _client_rows(rows)
        df = pd.DataFrame(compact_rows, columns=CLIENT_REVIEW_COLUMNS)
        df.to_csv(path, index=False, encoding="utf-8-sig", sep=";")
        _copy_assets_for_delivery(rows, path)
