from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from sendability import append_goldset_feedback


TRAINING_COLUMNS = [
    "company",
    "person",
    "role",
    "website",
    "current_line",
    "opening_line",
    "client_decision",
    "client_rewrite",
    "main_reason",
    "surface_to_focus_on",
    "evidence_or_context",
    "what_good_should_sound_like",
    "final_notes",
]

CLIENT_DECISIONS = ["Send as is", "Rewrite", "Reject"]

MAIN_REASONS = [
    "Good as is",
    "Tone wrong",
    "Too generic",
    "Unsupported claim",
    "Wrong surface",
    "Too technical",
    "Too long",
    "Missing outcome",
    "Signal-to-implication bridge",
    "Weak evidence",
    "Does not fit template",
    "Other",
]

SURFACES = [
    "App onboarding",
    "App Store / Google Play",
    "Review complaints",
    "Website landing page",
    "Booking flow",
    "Signup flow",
    "Checkout / paywall",
    "Case studies / proof",
    "Positioning",
    "Other",
]

REASON_TO_CATEGORY = {
    "good as is": "good_as_is",
    "tone wrong": "tone",
    "too generic": "too_generic",
    "unsupported claim": "unsupported_claim",
    "wrong surface": "wrong_surface",
    "too technical": "too_technical",
    "too long": "too_long",
    "missing outcome": "missing_outcome",
    "signal-to-implication bridge": "signal_to_implication_bridge",
    "weak evidence": "weak_evidence",
    "does not fit template": "bad_pitch_flow",
    "other": "other",
}

DECISION_TO_HUMAN = {
    "send as is": "send",
    "rewrite": "edit",
    "reject": "reject",
}

INSTRUCTIONS = [
    {
        "field": "opening_line",
        "what_it_means": "The actual opener the model wrote.",
        "how_to_fill": "Usually leave this as-is. Rewrite in client_rewrite instead so the original stays available for training.",
    },
    {
        "field": "client_decision",
        "what_it_means": "Choose Send as is, Rewrite, or Reject.",
        "how_to_fill": "Use Send as is only if you would actually send the line to a prospect.",
    },
    {
        "field": "client_rewrite",
        "what_it_means": "Your preferred version of the line.",
        "how_to_fill": "Only needed when client_decision is Rewrite. Keep it close to the tone you want.",
    },
    {
        "field": "main_reason",
        "what_it_means": "The main reason the line worked or failed.",
        "how_to_fill": "Pick the closest reason. If unsure, choose Other and explain in final_notes.",
    },
    {
        "field": "surface_to_focus_on",
        "what_it_means": "Where the personalization should have looked.",
        "how_to_fill": "For app-first products, this is usually app onboarding, app store, reviews, signup, or paywall.",
    },
    {
        "field": "evidence_or_context",
        "what_it_means": "The proof or context the line should use.",
        "how_to_fill": "Optional, but useful. Paste a short note, source URL, review theme, or app-flow observation.",
    },
    {
        "field": "what_good_should_sound_like",
        "what_it_means": "Plain-language tone guidance.",
        "how_to_fill": "Example: more conversational, less strategic, focus on current friction, shorter.",
    },
    {
        "field": "final_notes",
        "what_it_means": "Anything else that would help tune the model.",
        "how_to_fill": "Use this for nuance, examples, or client-specific preferences.",
    },
]

EXAMPLE_ROWS = [
    {
        "company": "Internal SDR Example",
        "person": "",
        "role": "",
        "website": "",
        "current_line": "Saw you are hiring SDRs.",
        "opening_line": "Saw you are hiring SDRs.",
        "client_decision": "Rewrite",
        "client_rewrite": "Saw you are hiring SDRs. Usually that means the team is trying to increase outbound volume without letting quality collapse.",
        "main_reason": "Signal-to-implication bridge",
        "surface_to_focus_on": "Other",
        "evidence_or_context": "Careers page lists SDR openings.",
        "what_good_should_sound_like": "Start with an observable signal, then connect it to a plausible business implication and operational tension.",
        "final_notes": "Pattern: signal_to_implication_bridge. Avoid signal-only lines, generic congratulations, unsupported assumptions, fake familiarity, or forcing multiple ideas.",
    },
    {
        "company": "Solo60",
        "person": "Ben",
        "role": "Founder",
        "website": "https://solo60.com",
        "current_line": "I was checking the website and noticed the location dropdown is not clearly clickable.",
        "opening_line": "I was checking the website and noticed the location dropdown is not clearly clickable.",
        "client_decision": "Rewrite",
        "client_rewrite": "I was checking the solo60 booking evidence and noticed the flow takes a few taps before available slots show, which could cost bookings from users looking to train today.",
        "main_reason": "Wrong surface",
        "surface_to_focus_on": "Booking flow",
        "evidence_or_context": "App-first product. Booking flow matters more than website dropdown.",
        "what_good_should_sound_like": "Specific, conversational, tied to bookings.",
        "final_notes": "Avoid pure UX nitpicks unless linked to a business outcome.",
    },
    {
        "company": "Hero",
        "person": "Brad",
        "role": "Founder",
        "website": "https://tryhero.app",
        "current_line": "The product is powerful but there is a lot introduced at once.",
        "opening_line": "The product is powerful but there is a lot introduced at once.",
        "client_decision": "Reject",
        "client_rewrite": "",
        "main_reason": "Too generic",
        "surface_to_focus_on": "App onboarding",
        "evidence_or_context": "First screen shows calendars, tasks, health, notes, and widgets.",
        "what_good_should_sound_like": "No generic praise. Say what a new user actually experiences.",
        "final_notes": "Focus on first-session churn or unclear starting point.",
    },
]


def review_df_to_training_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=TRAINING_COLUMNS)
    out = pd.DataFrame()
    out["company"] = df.get("company", "")
    out["person"] = df.get("person", "")
    out["role"] = df.get("role", "")
    out["website"] = df.get("website", "")
    out["current_line"] = df.get("template_preview", df.get("personalized_line", ""))
    out["opening_line"] = df.get("personalized_line", df.get("opening_line", ""))
    out["client_decision"] = ""
    out["client_rewrite"] = ""
    out["main_reason"] = ""
    out["surface_to_focus_on"] = df.get("surface_checked", "")
    out["evidence_or_context"] = df.get("evidence_found", "")
    out["what_good_should_sound_like"] = ""
    out["final_notes"] = ""
    return out[TRAINING_COLUMNS].fillna("")


def blank_training_df(row_count: int = 25) -> pd.DataFrame:
    return pd.DataFrame([{column: "" for column in TRAINING_COLUMNS} for _ in range(row_count)], columns=TRAINING_COLUMNS)


def _add_dropdown(ws, column_name: str, values: list[str], row_count: int) -> None:
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return
    col_idx = headers.index(column_name) + 1
    col_letter = ws.cell(1, col_idx).column_letter
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choose one of the dropdown values."
    validation.errorTitle = "Invalid value"
    ws.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{row_count + 1}")


def _style_training_workbook(buffer: BytesIO) -> bytes:
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Client Feedback"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    widths = {
        "company": 22,
        "person": 20,
        "role": 22,
        "website": 32,
        "current_line": 74,
        "opening_line": 74,
        "client_decision": 18,
        "client_rewrite": 80,
        "main_reason": 24,
        "surface_to_focus_on": 26,
        "evidence_or_context": 66,
        "what_good_should_sound_like": 52,
        "final_notes": 52,
    }
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[cell.column_letter].width = widths.get(str(cell.value), 24)
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 72
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    _add_dropdown(ws, "client_decision", CLIENT_DECISIONS, max(ws.max_row, 25))
    _add_dropdown(ws, "main_reason", MAIN_REASONS, max(ws.max_row, 25))
    _add_dropdown(ws, "surface_to_focus_on", SURFACES, max(ws.max_row, 25))

    instructions = wb["Instructions"]
    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
    instructions.column_dimensions["A"].width = 30
    instructions.column_dimensions["B"].width = 60
    instructions.column_dimensions["C"].width = 80
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    examples = wb["Examples"]
    for cell in examples[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        examples.column_dimensions[cell.column_letter].width = widths.get(str(cell.value), 24)
    for row in examples.iter_rows(min_row=2):
        examples.row_dimensions[row[0].row].height = 92
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def training_template_bytes(df: pd.DataFrame | None = None) -> bytes:
    training_df = review_df_to_training_df(df) if df is not None and not df.empty else blank_training_df()
    instructions_df = pd.DataFrame(INSTRUCTIONS)
    examples_df = pd.DataFrame(EXAMPLE_ROWS, columns=TRAINING_COLUMNS)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        training_df.to_excel(writer, index=False, sheet_name="Client Feedback")
        instructions_df.to_excel(writer, index=False, sheet_name="Instructions")
        examples_df.to_excel(writer, index=False, sheet_name="Examples")
    return _style_training_workbook(buffer)


def write_training_template(path: str | Path, df: pd.DataFrame | None = None) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(training_template_bytes(df))
    return output


def read_training_feedback(uploaded_or_path: Any) -> pd.DataFrame:
    if hasattr(uploaded_or_path, "getvalue"):
        name = str(getattr(uploaded_or_path, "name", "feedback.xlsx")).lower()
        data = BytesIO(uploaded_or_path.getvalue())
        if name.endswith(".csv"):
            return pd.read_csv(data, dtype=str).fillna("")
        return pd.read_excel(data, sheet_name="Client Feedback", dtype=str).fillna("")
    path = Path(uploaded_or_path)
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str).fillna("")
    return pd.read_excel(path, sheet_name="Client Feedback", dtype=str).fillna("")


def normalize_training_feedback(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy().fillna("")
    for column in TRAINING_COLUMNS:
        if column not in normalized:
            normalized[column] = ""
    normalized = normalized[TRAINING_COLUMNS]
    has_decision = normalized["client_decision"].astype(str).str.strip().ne("")
    normalized = normalized.loc[has_decision].copy()
    if normalized.empty:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["company"] = normalized["company"]
    out["person"] = normalized["person"]
    out["role"] = normalized["role"]
    out["website"] = normalized["website"]
    source_line = normalized["opening_line"].where(normalized["opening_line"].astype(str).str.strip().ne(""), normalized["current_line"])
    out["personalized_line"] = source_line
    out["model_opening_line"] = source_line
    out["current_opening_line"] = source_line
    out["template_preview"] = normalized["current_line"]
    out["human_decision"] = normalized["client_decision"].str.lower().map(DECISION_TO_HUMAN).fillna("unreviewed")
    out["edited_line"] = normalized["client_rewrite"]
    out["edit_reason_category"] = normalized["main_reason"].str.lower().map(REASON_TO_CATEGORY).fillna("other")
    out["edit_notes"] = normalized.apply(
        lambda row: " | ".join(
            item
            for item in [
                f"Surface: {row['surface_to_focus_on']}" if row["surface_to_focus_on"] else "",
                f"Good should sound like: {row['what_good_should_sound_like']}" if row["what_good_should_sound_like"] else "",
                row["final_notes"],
            ]
            if str(item).strip()
        ),
        axis=1,
    )
    out["evidence_found"] = normalized["evidence_or_context"]
    out["surface_checked"] = normalized["surface_to_focus_on"]
    out["source_urls"] = normalized["evidence_or_context"].where(normalized["evidence_or_context"].str.startswith("http"), "")
    out["status"] = "Ready"
    out["quality_flags"] = ""
    out["needs_manual_review"] = "no"
    out["visual_confidence"] = ""
    out["source_kind"] = "client_feedback"
    out["label_source"] = "client"
    out["origin_run_id"] = ""
    out["origin_row_id"] = ""
    out["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return out.fillna("")


def append_training_feedback_to_goldset(uploaded_or_path: Any, split: str = "candidate_training_set") -> tuple[Path, int, pd.DataFrame]:
    if split == "frozen_eval_set":
        raise ValueError("Client feedback imports cannot be saved directly to frozen_eval_set. Import to reviewed_examples or candidate_training_set first, then promote examples intentionally.")
    raw = read_training_feedback(uploaded_or_path)
    normalized = normalize_training_feedback(raw)
    path, count = append_goldset_feedback(normalized, split=split)
    return path, count, normalized
